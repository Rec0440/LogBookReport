"""
відкрити файл,
створити листок YearSummary-звіт,
додати в кінець workbook
"""

from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl import load_workbook
from datetime import datetime


class Report:
    # create workbook
    def __init__(self, filename, year):
        self.filename = filename                                # name of datafile
        self.default_year = year[2:]                            # two last digit of def year
        self.workb = load_workbook(filename)                    # data workbook
        self.ExecutionTime = datetime.now()                     # date and time when starting count
        self.ReportName = self.set_report_title()               # construct name of report
        self.workSh = self.workb.create_sheet(self.ReportName)  # create report worksheet at last position in workbook
        self.Name_Sheets = self.workb.sheetnames                # list of worksheets
        self.Name_Sheets.reverse()                              # reverse for counting from today into past
        self.length_sheets_names = len(self.Name_Sheets)        # number of worksheets

    def set_report_title(self):
        month = self.ExecutionTime.strftime('%m')
        my_title = 'Report20' + str(self.default_year) + '-' + month
        return my_title

    # create form of summary
    def create_workspace_page1(self, total_data):
        data = total_data

        # insert current time into A1, A2
        self.workSh['A1'].value = self.ExecutionTime.strftime('%d-%m-%y')
        self.workSh.merge_cells('A2:A3')
        self.workSh['A2'] = 'Year: 20' + self.default_year

        #    HEADER
        self.workSh.merge_cells('B1:J1')
        self.workSh['B1'] = 'PILOT RECORD/Облік польотів'
        self.workSh.merge_cells('B2:B3')
        self.workSh['B2'] = 'SINGLE-ENGINE'
        self.workSh.merge_cells('C2:C3')
        self.workSh['C2'] = 'MULTI-ENGINE'
        self.workSh.merge_cells('D2:D3')
        self.workSh['D2'] = 'MULTI-ENGINE MULTI-PILOT'
        self.workSh.merge_cells('E2:E3')
        self.workSh['E2'].value = 'JET'
        self.workSh.merge_cells('F2:F3')
        self.workSh['F2'].value = 'TURBOPROP'

        self.workSh.column_dimensions['A'].width = 16
        self.workSh.column_dimensions['B'].width = 11.0
        self.workSh.column_dimensions['C'].width = 11.0
        self.workSh.column_dimensions['D'].width = 14
        self.workSh.column_dimensions['E'].width = 7.0
        self.workSh.column_dimensions['F'].width = 12

        self.workSh.merge_cells('G2:G3')
        self.workSh['G2'] = 'TOTAL FLIGHT TIME'
        self.workSh.merge_cells('H2:I2')
        self.workSh['H2'].value = 'LANDINGS'
        self.workSh['H3'].value = 'DAY'
        self.workSh['I3'].value = 'NIGHT'
        self.workSh.merge_cells('J2:J3')
        self.workSh['J2'] = 'LANDINGS TOTAL'
        self.workSh.column_dimensions['G'].width = 14
        self.workSh.column_dimensions['H'].width = 5
        self.workSh.column_dimensions['I'].width = 7
        self.workSh.column_dimensions['J'].width = 10

        # main body
        cel_list = ['A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A11', 'A12', 'A13', 'A14', 'A15', 'A16']
        months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
        for cel, mon in zip(cel_list, months):
            self.workSh[cel].value = data[mon]['month']

        se_list = ['B'+cel[1:] for cel in cel_list]
        for cel, mon in zip(se_list, months):
            self.workSh[cel] = data[mon]['flight_time']['single_engine']['sum']

        me_list = ['C'+cel[1:] for cel in cel_list]
        for cel, mon in zip(me_list, months):
            self.workSh[cel] = data[mon]['flight_time']['multi_engine']['sum']

        mp_list = ['D'+cel[1:] for cel in cel_list]
        for cel, mon in zip(mp_list, months):
            self.workSh[cel] = data[mon]['flight_time']['multi_pilot']['sum']

        tot_time_list = ['G'+cel[1:] for cel in cel_list]
        for cel, mon in zip(tot_time_list, months):
            self.workSh[cel] = data[mon]['flight_time']['total_flight_time']['sum']

        dland_list = ['H' + cel[1:] for cel in cel_list]
        for cel, mon in zip(dland_list, months):
            dayland = data[mon]['landings']['day']
            self.workSh[cel].value = sum(x if x is not None else 0 for x in dayland)

        nland_list = ['I' + cel[1:] for cel in cel_list]
        for cel, mon in zip(nland_list, months):
            nightland = data[mon]['landings']['night']
            self.workSh[cel].value = sum(x if x is not None else 0 for x in nightland)

        tot_land_list = ['J' + cel[1:] for cel in cel_list]
        for cel, mon in zip(tot_land_list, months):
            self.workSh[cel].value = data[mon]['landings']['total'] if data[mon]['landings']['total'] else 0

        self.workSh['A10'].value = 'TOTALS 6mo'
        self.workSh['B10'].value = data['total_6mo']['single_engine']
        self.workSh['C10'].value = data['total_6mo']['multi_engine']
        self.workSh['D10'].value = data['total_6mo']['multi_pilot']
        self.workSh['G10'].value = data['total_6mo']['total_flight_time']
        self.workSh['H10'].value = data['total_6mo']['day_land']
        self.workSh['I10'].value = data['total_6mo']['night_land']
        self.workSh['J10'].value = data['total_6mo']['total_land']

        self.workSh['A17'].value = 'YEAR TOTALS'
        self.workSh['B17'].value = data['total_year']['single_engine']
        self.workSh['C17'].value = data['total_year']['multi_engine']
        self.workSh['D17'].value = data['total_year']['multi_pilot']
        self.workSh['G17'].value = data['total_year']['total_flight_time']
        self.workSh['H17'].value = data['total_year']['day_land']
        self.workSh['I17'].value = data['total_year']['night_land']
        self.workSh['J17'].value = data['total_year']['total_land']

        self.workSh['A18'].value = 'TOTALS prev.YEARS'
        self.workSh['A19'].value = 'TOTALS TO DATE'

    # create second part of form of summary
    # ____________________!!!!!!!!!!!______________________
    def create_workspace_page2(self, total_data):
        data = total_data

        # Style preparing
        ft = Font(bold=True)
        bd = Side(style='medium')
        set_border = Border(left=bd, bottom=bd, right=bd, top=bd)
        hv_alignment = Alignment(horizontal='center', vertical='center')
        hvw_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hvw_array = ['B2', 'C2', 'D2', 'F2', 'G2', 'J2', 'K1', 'O1', 'P1', 'Q1', 'S2', 'U2']

        #    HEADER
        self.workSh.merge_cells('K1:L2')
        self.workSh['K1'] = 'OPERATIONAL CONDITION TIME'
        self.workSh['K3'] = 'NIGHT'
        self.workSh['L3'] = 'IFR'
        self.workSh.merge_cells('M1:M3')
        self.workSh['M1'] = 'PIC'
        self.workSh.column_dimensions['K'].width = 10
        self.workSh.column_dimensions['L'].width = 10
        self.workSh.column_dimensions['M'].width = 10

        self.workSh.merge_cells('N1:N3')
        self.workSh['N1'] = 'CO-PILOT'
        self.workSh.merge_cells('O1:O3')
        self.workSh['O1'] = 'DUAL RECEIVED'
        self.workSh.merge_cells('P1:P3')
        self.workSh['P1'] = 'INSTRUCTOR'
        self.workSh.merge_cells('Q1:Q3')
        self.workSh['Q1'] = 'FLIGHT SIMULATOR'
        self.workSh.column_dimensions['N'].width = 11
        self.workSh.column_dimensions['O'].width = 11
        self.workSh.column_dimensions['P'].width = 13
        self.workSh.column_dimensions['Q'].width = 11

        self.workSh.merge_cells('R1:U1')
        self.workSh['R1'] = 'INSTRUCTION GIVEN'
        self.workSh.merge_cells('R2:R3')
        self.workSh['R2'] = 'TOTAL'
        self.workSh.merge_cells('S2:S3')
        self.workSh['S2'] = 'IFR/ NIGHT'
        self.workSh.merge_cells('T2:T3')
        self.workSh['T2'] = 'ME'
        self.workSh.merge_cells('U2:U3')
        self.workSh['U2'] = 'PRIVAT/ COMMERCIAL'
        self.workSh['U2'].font = 'PRIVAT/ COMMERCIAL'

        self.workSh.column_dimensions['R'].width = 9
        self.workSh.column_dimensions['S'].width = 9
        self.workSh.column_dimensions['T'].width = 9
        self.workSh.column_dimensions['U'].width = 9

        # main body
        cel_list = ['K4', 'K5', 'K6', 'K7', 'K8', 'K9', 'K11', 'K12', 'K13', 'K14', 'K15', 'K16']
        months = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
        for cel, mon in zip(cel_list, months):
            self.workSh[cel] = data[mon]['flight_time']['night_cond']['sum']

        pic_list = ['M'+cel[1:] for cel in cel_list]
        for cel, mon in zip(pic_list, months):
            self.workSh[cel] = data[mon]['flight_time']['PIC']['sum']

        copilot_list = ['N'+cel[1:] for cel in cel_list]
        for cel, mon in zip(copilot_list, months):
            self.workSh[cel] = data[mon]['flight_time']['Co-Pilot']['sum']

        dual_list = ['O'+cel[1:] for cel in cel_list]
        for cel, mon in zip(dual_list, months):
            self.workSh[cel] = data[mon]['flight_time']['dual']['sum']

        instr_list = ['P'+cel[1:] for cel in cel_list]
        for cel, mon in zip(instr_list, months):
            self.workSh[cel] = data[mon]['flight_time']['INSTRUCTOR']['sum']

        self.workSh['K10'].value = data['total_6mo']['night_cond']
        self.workSh['L10'].value = data['total_6mo']['ifr_cond']
        self.workSh['M10'].value = data['total_6mo']['PIC']
        self.workSh['N10'].value = data['total_6mo']['Co-Pilot']
        self.workSh['O10'].value = data['total_6mo']['dual']
        self.workSh['P10'].value = data['total_6mo']['INSTRUCTOR']
        self.workSh['K17'].value = data['total_year']['night_cond']
        self.workSh['L17'].value = data['total_year']['ifr_cond']
        self.workSh['M17'].value = data['total_year']['PIC']
        self.workSh['N17'].value = data['total_year']['Co-Pilot']
        self.workSh['O17'].value = data['total_year']['dual']
        self.workSh['P17'].value = data['total_year']['INSTRUCTOR']

        # set styles and alignment
        for col in self.workSh.columns:
            for cel in col:
                cel.alignment = hv_alignment
                cel.font = ft
                cel.border = set_border
        for cel in hvw_array:
            self.workSh[cel].alignment = hvw_alignment

    # prepare print options
    def set_print_option(self):
        self.workSh.set_printer_settings('9', 'landscape')

    # save result of calculating
    def save_summaries(self):
        self.workb.save(self.filename)
